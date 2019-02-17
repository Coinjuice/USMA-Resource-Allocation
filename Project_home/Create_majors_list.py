#! /usr/bin/env python
#./filter_majors.py fos_maj_data.xlsx fos B A D F
import openpyxl as px
import sys
import csv
import subprocess
import os

def make_major_dict(filename,sheetname,majorColumn,courseColumn,prereqColumn,filterColumn):
	majorColumn = str(majorColumn).upper()
	courseColumn = str(courseColumn).upper()
	wb = px.load_workbook(filename)
	inf = wb.active
	p = wb.get_sheet_by_name(name=sheetname)

	os.system("mkdir majors")
	maj_outf = open("majors/"+inf[majorColumn+str(1)].value+".txt","a")
	current_major = inf[majorColumn+str(1)].value
	cdoc=current_major+".csv"
	f=open("majors/"+cdoc,"a")
	writer = csv.writer(f)
	for i in range(1,p.max_row+1):
		

		if inf[majorColumn+str(int(i))].value!=current_major:
			maj_outf.close()
			maj_outf = open("majors/"+inf[majorColumn+str(int(i))].value+".txt","a")
			current_major = inf[majorColumn+str(int(i))].value
			cdoc=current_major+".csv"
			f=open("majors/"+cdoc,"a")
			writer = csv.writer(f)
		
		print("process: "+str(i)+"/"+str(p.max_row+1)+"\n\n")
		if inf[filterColumn+str(int(i))].value=="PRE REQUISITE":
				
			edge_list = set()
			prereq= str(inf[prereqColumn+str(int(i))].value)
			course = str(inf[courseColumn+str(int(i))].value)
			edge_list.add(prereq+";"+course)
			writer.writerow([prereq.replace(" ","")+";"+course])
	# 		inf = open("majors/"+doc) Don't want to reopen this file and read from a string use the dictionary
	#		for line in inf:
	#			writer.writerow([line])
			
			to_insert=str(edge_list).replace(" ","").replace("'","").replace("[","").replace("]","").replace(",","\n").replace("crse_nbr;rltd_crse_nbr\n","").replace("set(","").replace(")","")
			maj_outf.write(to_insert)
			
		
	#		inf.close()
	f.close()
	maj_outf.close()
	return majors
if __name__=='__main__':
	inf = sys.argv[1]
	sheet = sys.argv[2]
	major = sys.argv[3]
	course = sys.argv[4]
	prereq = sys.argv[5]
	filterMethod = sys.argv[6]
	make_major_dict(inf,sheet,major,course,prereq,filterMethod)
